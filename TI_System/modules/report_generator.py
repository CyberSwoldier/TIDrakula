#!/usr/bin/env python3
"""
Report Generator Module for Threat Intelligence System
File: modules/report_generator.py
"""
import pandas as pd
import json
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Any
from pathlib import Path
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Generates comprehensive threat intelligence reports in various formats"""
    
    def __init__(self, config: dict = None):
        """
        Initialize report generator
        
        Args:
            config: Configuration dictionary
        """
        self.config = config or {}
        self._setup_styles()
    
    def _setup_styles(self):
        """Setup Excel styles for professional reports"""
        self.styles = {
            'header': {
                'font': Font(bold=True, color='FFFFFF', size=12),
                'fill': PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'data': {
                'font': Font(size=10),
                'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'title': {
                'font': Font(bold=True, size=16, color='2E86AB'),
                'alignment': Alignment(horizontal='center')
            },
            'summary': {
                'font': Font(bold=True, size=11, color='D2691E'),
                'fill': PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid'),
                'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            }
        }
    
    def generate_excel_report(self, human_df: pd.DataFrame, general_df: pd.DataFrame, 
                            output_file: str, include_summary: bool = True) -> bool:
        """
        Generate comprehensive Excel report with multiple sheets
        
        Args:
            human_df: DataFrame with human-targeted attacks
            general_df: DataFrame with general attacks
            output_file: Output file path
            include_summary: Whether to include summary sheet
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Process dataframes
            human_processed = self._process_dataframe_for_excel(human_df)
            general_processed = self._process_dataframe_for_excel(general_df)
            
            # Create Excel writer
            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                # Write data sheets
                if not human_processed.empty:
                    human_processed.to_excel(writer, sheet_name="Human_Attacks", index=False)
                    self._format_sheet(writer.sheets["Human_Attacks"], "Human-Targeted Attacks")
                
                if not general_processed.empty:
                    general_processed.to_excel(writer, sheet_name="General_Attacks", index=False)
                    self._format_sheet(writer.sheets["General_Attacks"], "General Attacks")
                
                # Generate summary sheet
                if include_summary:
                    summary_data = self._create_summary_data(human_df, general_df)
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name="Executive_Summary", index=False)
                    self._format_summary_sheet(writer.sheets["Executive_Summary"])
                
                # Generate statistics sheet
                stats_data = self._create_statistics_data(human_df, general_df)
                stats_df = pd.DataFrame(stats_data)
                stats_df.to_excel(writer, sheet_name="Statistics", index=False)
                self._format_sheet(writer.sheets["Statistics"], "Threat Intelligence Statistics")
            
            logger.info(f"Excel report generated successfully: {output_file}")
            return True
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            return False
    
    def _process_dataframe_for_excel(self, df: pd.DataFrame) -> pd.DataFrame:
        """Process dataframe for Excel export"""
        if df.empty:
            return df
        
        try:
            # Create a copy to avoid modifying original
            processed_df = df.copy()
            
            # Split TTPs into separate columns if they exist as comma-separated
            if 'ttps' in processed_df.columns:
                ttps_split = processed_df['ttps'].str.split(',', expand=True)
                for i, col in enumerate(ttps_split.columns):
                    if col is not None:
                        processed_df[f'ttp_desc_{i+1}'] = ttps_split[col].str.strip() if ttps_split[col] is not None else None
                processed_df = processed_df.drop('ttps', axis=1)
            
            # Split countries into separate columns if they exist as comma-separated
            if 'countries' in processed_df.columns:
                countries_split = processed_df['countries'].str.split(',', expand=True)
                for i, col in enumerate(countries_split.columns):
                    if col is not None:
                        processed_df[f'country_{i+1}'] = countries_split[col].str.strip() if countries_split[col] is not None else None
                processed_df = processed_df.drop('countries', axis=1)
            
            # Format datetime columns
            for col in processed_df.columns:
                if 'date' in col.lower() or 'time' in col.lower():
                    processed_df[col] = pd.to_datetime(processed_df[col], errors='coerce')
            
            # Clean up text columns
            text_columns = ['title', 'content', 'source', 'threat_actor']
            for col in text_columns:
                if col in processed_df.columns:
                    processed_df[col] = processed_df[col].astype(str).str.strip()
                    processed_df[col] = processed_df[col].replace('nan', '')
            
            # Reorder columns for better presentation
            desired_order = [
                'id', 'title', 'source', 'published_date', 'threat_actor', 
                'severity', 'is_human_targeted'
            ]
            
            # Add TTP columns
            ttp_cols = [col for col in processed_df.columns if col.startswith('ttp_desc_')]
            desired_order.extend(sorted(ttp_cols))
            
            # Add country columns
            country_cols = [col for col in processed_df.columns if col.startswith('country_')]
            desired_order.extend(sorted(country_cols))
            
            # Add remaining columns
            remaining_cols = [col for col in processed_df.columns if col not in desired_order]
            desired_order.extend(remaining_cols)
            
            # Filter out columns that don't exist
            final_order = [col for col in desired_order if col in processed_df.columns]
            
            return processed_df[final_order]
            
        except Exception as e:
            logger.error(f"Error processing dataframe for Excel: {e}")
            return df
    
    def _format_sheet(self, worksheet, title: str):
        """Format Excel worksheet with professional styling"""
        try:
            # Add title
            worksheet.insert_rows(1, 2)
            worksheet['A1'] = title
            worksheet['A1'].font = self.styles['title']['font']
            worksheet['A1'].alignment = self.styles['title']['alignment']
            worksheet.merge_cells('A1:Z1')
            
            # Format headers (now in row 3)
            header_row = 3
            for cell in worksheet[header_row]:
                if cell.value:
                    cell.font = self.styles['header']['font']
                    cell.fill = self.styles['header']['fill']
                    cell.alignment = self.styles['header']['alignment']
                    cell.border = self.styles['header']['border']
            
            # Format data rows
            for row_num in range(4, worksheet.max_row + 1):
                for cell in worksheet[row_num]:
                    cell.font = self.styles['data']['font']
                    cell.alignment = self.styles['data']['alignment']
                    cell.border = self.styles['data']['border']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                # Set reasonable width limits
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = max(adjusted_width, 10)
            
            # Freeze panes at header row
            worksheet.freeze_panes = 'A4'
            
        except Exception as e:
            logger.error(f"Error formatting Excel sheet: {e}")
    
    def _format_summary_sheet(self, worksheet):
        """Format summary sheet with special styling"""
        try:
            # Format all cells as summary style
            for row in worksheet:
                for cell in row:
                    if cell.value:
                        cell.font = self.styles['summary']['font']
                        cell.fill = self.styles['summary']['fill']
                        cell.alignment = self.styles['summary']['alignment']
                        cell.border = self.styles['summary']['border']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = min(max_length + 2, 80)
                worksheet.column_dimensions[column_letter].width = max(adjusted_width, 15)
                
        except Exception as e:
            logger.error(f"Error formatting summary sheet: {e}")
    
    def _create_summary_data(self, human_df: pd.DataFrame, general_df: pd.DataFrame) -> List[Dict]:
        """Create executive summary data"""
        try:
            total_threats = len(human_df) + len(general_df)
            human_threats = len(human_df)
            general_threats = len(general_df)
            
            # Calculate time period
            all_dates = []
            if not human_df.empty and 'published_date' in human_df.columns:
                all_dates.extend(pd.to_datetime(human_df['published_date'], errors='coerce').dropna())
            if not general_df.empty and 'published_date' in general_df.columns:
                all_dates.extend(pd.to_datetime(general_df['published_date'], errors='coerce').dropna())
            
            if all_dates:
                start_date = min(all_dates).strftime('%Y-%m-%d')
                end_date = max(all_dates).strftime('%Y-%m-%d')
                date_range = f"{start_date} to {end_date}"
            else:
                date_range = "Unknown"
            
            # Count unique sources
            sources = set()
            if not human_df.empty and 'source' in human_df.columns:
                sources.update(human_df['source'].dropna().unique())
            if not general_df.empty and 'source' in general_df.columns:
                sources.update(general_df['source'].dropna().unique())
            
            # Count unique countries
            countries = set()
            if not human_df.empty and 'countries' in human_df.columns:
                for countries_str in human_df['countries'].dropna():
                    if isinstance(countries_str, str):
                        countries.update([c.strip() for c in countries_str.split(',')])
            if not general_df.empty and 'countries' in general_df.columns:
                for countries_str in general_df['countries'].dropna():
                    if isinstance(countries_str, str):
                        countries.update([c.strip() for c in countries_str.split(',')])
            
            # Count unique TTPs
            ttps = set()
            if not human_df.empty and 'ttps' in human_df.columns:
                for ttps_str in human_df['ttps'].dropna():
                    if isinstance(ttps_str, str):
                        ttps.update([t.strip() for t in ttps_str.split(',')])
            if not general_df.empty and 'ttps' in general_df.columns:
                for ttps_str in general_df['ttps'].dropna():
                    if isinstance(ttps_str, str):
                        ttps.update([t.strip() for t in ttps_str.split(',')])
            
            summary_data = [
                {"Metric": "Report Generation Date", "Value": datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
                {"Metric": "Analysis Period", "Value": date_range},
                {"Metric": "Total Threats Analyzed", "Value": str(total_threats)},
                {"Metric": "Human-Targeted Attacks", "Value": str(human_threats)},
                {"Metric": "General Attacks", "Value": str(general_threats)},
                {"Metric": "Human-Targeted Percentage", "Value": f"{(human_threats/total_threats*100):.1f}%" if total_threats > 0 else "0%"},
                {"Metric": "Unique Sources", "Value": str(len(sources))},
                {"Metric": "Affected Countries", "Value": str(len(countries))},
                {"Metric": "Attack Techniques (TTPs)", "Value": str(len(ttps))},
            ]
            
            return summary_data
            
        except Exception as e:
            logger.error(f"Error creating summary data: {e}")
            return [{"Metric": "Error", "Value": "Could not generate summary"}]
    
    def _create_statistics_data(self, human_df: pd.DataFrame, general_df: pd.DataFrame) -> List[Dict]:
        """Create detailed statistics data"""
        try:
            stats_data = []
            
            # Top sources
            all_sources = []
            if not human_df.empty and 'source' in human_df.columns:
                all_sources.extend(human_df['source'].dropna().tolist())
            if not general_df.empty and 'source' in general_df.columns:
                all_sources.extend(general_df['source'].dropna().tolist())
            
            if all_sources:
                from collections import Counter
                top_sources = Counter(all_sources).most_common(10)
                stats_data.extend([
                    {"Category": "Top Sources", "Item": source, "Count": count}
                    for source, count in top_sources
                ])
            
            # Top TTPs
            all_ttps = []
            for df in [human_df, general_df]:
                if not df.empty and 'ttps' in df.columns:
                    for ttps_str in df['ttps'].dropna():
                        if isinstance(ttps_str, str):
                            all_ttps.extend([t.strip() for t in ttps_str.split(',')])
            
            if all_ttps:
                from collections import Counter
                top_ttps = Counter(all_ttps).most_common(10)
                stats_data.extend([
                    {"Category": "Top TTPs", "Item": ttp, "Count": count}
                    for ttp, count in top_ttps
                ])
            
            # Top countries
            all_countries = []
            for df in [human_df, general_df]:
                if not df.empty and 'countries' in df.columns:
                    for countries_str in df['countries'].dropna():
                        if isinstance(countries_str, str):
                            all_countries.extend([c.strip() for c in countries_str.split(',')])
            
            if all_countries:
                from collections import Counter
                top_countries = Counter(all_countries).most_common(10)
                stats_data.extend([
                    {"Category": "Top Countries", "Item": country, "Count": count}
                    for country, count in top_countries
                ])
            
            return stats_data
            
        except Exception as e:
            logger.error(f"Error creating statistics data: {e}")
            return [{"Category": "Error", "Item": "Could not generate statistics", "Count": 0}]
    
    def generate_json_report(self, human_df: pd.DataFrame, general_df: pd.DataFrame, 
                           output_file: str) -> bool:
        """
        Generate JSON report
        
        Args:
            human_df: DataFrame with human-targeted attacks
            general_df: DataFrame with general attacks  
            output_file: Output file path
            
        Returns:
            True if successful, False otherwise
        """
        try:
            report_data = {
                "metadata": {
                    "generated_at": datetime.now().isoformat(),
                    "report_type": "threat_intelligence",
                    "version": "1.0"
                },
                "summary": self._create_summary_data(human_df, general_df),
                "statistics": self._create_statistics_data(human_df, general_df),
                "human_targeted_attacks": human_df.to_dict('records') if not human_df.empty else [],
                "general_attacks": general_df.to_dict('records') if not general_df.empty else []
            }
            
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=2, default=str, ensure_ascii=False)
            
            logger.info(f"JSON report generated successfully: {output_file}")
            return True
            
        except Exception as e:
            logger.error(f"Error generating JSON report: {e}")
            return False
    
    def generate_csv_report(self, human_df: pd.DataFrame, general_df: pd.DataFrame, 
                          output_dir: str) -> bool:
        """
        Generate CSV reports (separate files for each category)
        
        Args:
            human_df: DataFrame with human-targeted attacks
            general_df: DataFrame with general attacks
            output_dir: Output directory path
            
        Returns:
            True if successful, False otherwise
        """
        try:
            output_path = Path(output_dir)
            output_path.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # Generate human-targeted attacks CSV
            if not human_df.empty:
                human_file = output_path / f"human_targeted_attacks_{timestamp}.csv"
                processed_human = self._process_dataframe_for_excel(human_df)
                processed_human.to_csv(human_file, index=False, encoding='utf-8')
                logger.info(f"Human-targeted attacks CSV saved: {human_file}")
            
            # Generate general attacks CSV
            if not general_df.empty:
                general_file = output_path / f"general_attacks_{timestamp}.csv"
                processed_general = self._process_dataframe_for_excel(general_df)
                processed_general.to_csv(general_file, index=False, encoding='utf-8')
                logger.info(f"General attacks CSV saved: {general_file}")
            
            # Generate summary CSV
            summary_file = output_path / f"executive_summary_{timestamp}.csv"
            summary_data = self._create_summary_data(human_df, general_df)
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_csv(summary_file, index=False, encoding='utf-8')
            logger.info(f"Executive summary CSV saved: {summary_file}")
            
            return True
            
        except Exception as e:
            logger.error(f"Error generating CSV reports: {e}")
            return False
    
    def generate_all_formats(self, human_df: pd.DataFrame, general_df: pd.DataFrame, 
                           base_filename: str) -> Dict[str, bool]:
        """
        Generate reports in all supported formats
        
        Args:
            human_df: DataFrame with human-targeted attacks
            general_df: DataFrame with general attacks
            base_filename: Base filename (without extension)
            
        Returns:
            Dictionary with format success status
        """
        results = {}
        
        try:
            # Excel report
            excel_file = f"{base_filename}.xlsx"
            results['excel'] = self.generate_excel_report(human_df, general_df, excel_file)
            
            # JSON report  
            json_file = f"{base_filename}.json"
            results['json'] = self.generate_json_report(human_df, general_df, json_file)
            
            # CSV reports
            csv_dir = f"{base_filename}_csv"
            results['csv'] = self.generate_csv_report(human_df, general_df, csv_dir)
            
            logger.info(f"Generated reports with results: {results}")
            return results
            
        except Exception as e:
            logger.error(f"Error generating all format reports: {e}")
            return {format_type: False for format_type in ['excel', 'json', 'csv']}
    
    def create_download_buffer(self, human_df: pd.DataFrame, general_df: pd.DataFrame, 
                             format_type: str = 'excel') -> Optional[BytesIO]:
        """
        Create a download buffer for the specified format
        
        Args:
            human_df: DataFrame with human-targeted attacks
            general_df: DataFrame with general attacks
            format_type: Format type ('excel', 'csv', 'json')
            
        Returns:
            BytesIO buffer or None if error
        """
        try:
            buffer = BytesIO()
            
            if format_type.lower() == 'excel':
                # Create Excel in memory
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    if not human_df.empty:
                        human_processed = self._process_dataframe_for_excel(human_df)
                        human_processed.to_excel(writer, sheet_name="Human_Attacks", index=False)
                    
                    if not general_df.empty:
                        general_processed = self._process_dataframe_for_excel(general_df)
                        general_processed.to_excel(writer, sheet_name="General_Attacks", index=False)
                    
                    # Add summary
                    summary_data = self._create_summary_data(human_df, general_df)
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name="Executive_Summary", index=False)
            
            elif format_type.lower() == 'csv':
                # Create combined CSV
                combined_df = pd.concat([human_df, general_df], ignore_index=True)
                processed_combined = self._process_dataframe_for_excel(combined_df)
                processed_combined.to_csv(buffer, index=False, encoding='utf-8')
            
            elif format_type.lower() == 'json':
                # Create JSON
                report_data = {
                    "metadata": {
                        "generated_at": datetime.now().isoformat(),
                        "report_type": "threat_intelligence",
                        "version": "1.0"
                    },
                    "human_targeted_attacks": human_df.to_dict('records') if not human_df.empty else [],
                    "general_attacks": general_df.to_dict('records') if not general_df.empty else []
                }
                
                json_str = json.dumps(report_data, indent=2, default=str, ensure_ascii=False)
                buffer.write(json_str.encode('utf-8'))
            
            buffer.seek(0)
            return buffer
            
        except Exception as e:
            logger.error(f"Error creating download buffer: {e}")
            return None
